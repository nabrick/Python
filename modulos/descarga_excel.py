import os
import math
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

def obtener_nombre_disponible(nombre_archivo):
    """
    Si el archivo existe, genera Archivo_1.xlsx, Archivo_2.xlsx, etc.
    """
    base, ext = os.path.splitext(nombre_archivo)
    contador = 1
    nuevo_nombre = nombre_archivo

    while os.path.exists(nuevo_nombre):
        nuevo_nombre = f"{base}_{contador}{ext}"
        contador += 1

    return nuevo_nombre

def exportar_excel(df, nombre_archivo='Archivo.xlsx', nombre_hoja='Reporte', agregar_hoja=False):
    """
    Exporta un DataFrame a Excel.
    Agregar_hoja = False:
        Si el archivo existe, crea uno nuevo agregando _1, _2, etc.
    Agregar_hoja = True:
        Si el archivo existe, agrega nuevas hojas al mismo archivo.
    """

    max_filas = 1_000_000
    total_filas = len(df)
    num_hojas = math.ceil(total_filas / max_filas)

    archivo_existe = os.path.exists(nombre_archivo)

    if agregar_hoja:
        mode = 'a' if archivo_existe else 'w'
    else:
        nombre_archivo = obtener_nombre_disponible(nombre_archivo)
        mode = 'w'

    with pd.ExcelWriter(
        nombre_archivo,
        engine='openpyxl',
        mode=mode,
        if_sheet_exists='new'
    ) as writer:

        for i in range(num_hojas):
            inicio = i * max_filas
            fin = inicio + max_filas
            df_parcial = df.iloc[inicio:fin]

            nombre_hoja_actual = (
                f"{nombre_hoja}_{i+1}" if num_hojas > 1 else nombre_hoja
            )

            df_parcial.to_excel(
                writer,
                sheet_name=nombre_hoja_actual,
                index=False
            )

            ws = writer.book[nombre_hoja_actual]

            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(
                start_color='4F81BD',
                end_color='4F81BD',
                fill_type='solid'
            )
            center_alignment = Alignment(horizontal='center', vertical='center')

            for col_num, cell in enumerate(ws[1], 1):
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment

                col_letter = get_column_letter(col_num)
                max_length = max(
                    len(str(c.value)) if c.value is not None else 0
                    for c in ws[col_letter]
                )
                ws.column_dimensions[col_letter].width = max_length + 2

    print(f"Archivo generado correctamente: {nombre_archivo}")